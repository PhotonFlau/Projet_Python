# -*- coding: utf-8 -*-
"""
Created on Tue Feb 27 11:52:17 2024

@author: Florian
"""

# =============================================================================
# Bibliothèques
import openpyxl as op
import matplotlib.pyplot as plt
# =============================================================================


# =============================================================================
# Fonctions
# =============================================================================

#Aluminium.....................................................................
def Affiche_Aluminium():
    
    book = op.load_workbook('Données.xlsx')
    
    #Variables
    Aluminium = book.get_sheet_by_name("Aluminium")
    
    #Listes
    energie = []
    tau = []
    tauPA = []
    tauNP = []
    tauEP = []
    
    for i in range(4,Aluminium.max_row+1):
        energie.append(float(Aluminium.cell(row=i, column=1).value))
        tau.append(float(Aluminium.cell(row=i, column=2).value))
        tauPA.append(float(Aluminium.cell(row=i, column=3).value))
        tauNP.append(float(Aluminium.cell(row=i, column=4).value))
        tauEP.append(float(Aluminium.cell(row=i, column=5).value))
    
    #Vérification
    print("\nListe tau : ", tau)
    print("\nListe tauPA : ", tauPA)
    print("\nListe tauNP : ", tauNP)
    print("\nListe tauEP : ", tauEP)
    
    plt.xlabel("Energie du photon (MEV)")
    plt.ylabel('Tau (cm2/g)')
    
    plt.xlim(0.001,1e+05)
    plt.ylim(1e-09,1e+04)
    
    plt.xscale('log')
    plt.yscale('log')
    
    plt.grid(True,which="both",linestyle='--')
    
    
    plt.plot(energie, tau, linewidth=1, label = 'Tau Inc.Scatter')
    plt.plot(energie, tauPA, linewidth=1, label = 'Tau Photoel.Abs')
    plt.plot(energie, tauNP, linewidth=1, label = 'Tau Nuclear.Pr')
    plt.plot(energie, tauEP, linewidth=1, label = 'Tau Elec.Pr')
    
    plt.legend()

    book.save("Données.xlsx")
    
#Plomb.........................................................................
def Affiche_Plomb():
    
    book = op.load_workbook('Données.xlsx')
    
    #Variables
    Plomb = book.get_sheet_by_name("Plomb")
    
    #Listes
    energie = []
    tau = []
    tauPA = []
    tauNP = []
    tauEP = []
    
    for i in range(4,Plomb.max_row+1):
        energie.append(float(Plomb.cell(row=i, column=1).value))
        tau.append(float(Plomb.cell(row=i, column=2).value))
        tauPA.append(float(Plomb.cell(row=i, column=3).value))
        tauNP.append(float(Plomb.cell(row=i, column=4).value))
        tauEP.append(float(Plomb.cell(row=i, column=5).value))
    
    #Vérification
    print("\nListe tau : ", tau)
    print("\nListe tauPA : ", tauPA)
    print("\nListe tauNP : ", tauNP)
    print("\nListe tauEP : ", tauEP)
    
    plt.xlabel("Energie du photon (MEV)")
    plt.ylabel('Tau (cm2/g)')
    
    plt.xlim(0.001,1e+05)
    plt.ylim(1e-09,1e+04)
    
    plt.xscale('log')
    plt.yscale('log')
    
    plt.grid(True,which="both",linestyle='--')
    
    
    plt.plot(energie, tau, linewidth=1, label = 'Tau Inc.Scatter')
    plt.plot(energie, tauPA, linewidth=1, label = 'Tau Photoel.Abs')
    plt.plot(energie, tauNP, linewidth=1, label = 'Tau Nuclear.Pr')
    plt.plot(energie, tauEP, linewidth=1, label = 'Tau Elec.Pr')
    
    plt.legend()

    book.save("Données.xlsx")
    
#Cobalt........................................................................
def Affiche_Cobalt():
    
    book = op.load_workbook('Données.xlsx')
    
    #Variables
    Cobalt = book['Cobalt']
    
    #Listes
    energie = []
    tau = []
    tauPA = []
    tauNP = []
    tauEP = []
    
    for i in range(4,Cobalt.max_row+1):
        energie.append(float(Cobalt.cell(row=i, column=1).value))
        tau.append(float(Cobalt.cell(row=i, column=2).value))
        tauPA.append(float(Cobalt.cell(row=i, column=3).value))
        tauNP.append(float(Cobalt.cell(row=i, column=4).value))
        tauEP.append(float(Cobalt.cell(row=i, column=5).value))
    
    #Vérification
    print("\nListe tau : ", tau)
    print("\nListe tauPA : ", tauPA)
    print("\nListe tauNP : ", tauNP)
    print("\nListe tauEP : ", tauEP)
    
    plt.xlabel("Energie du photon (MEV)")
    plt.ylabel('Tau (cm2/g)')
    
    plt.xlim(0.001,1e+05)
    plt.ylim(1e-09,1e+04)
    
    plt.xscale('log')
    plt.yscale('log')
    
    plt.grid(True,which="both",linestyle='--')
    
    
    plt.plot(energie, tau, linewidth=1, label = 'Tau Inc.Scatter')
    plt.plot(energie, tauPA, linewidth=1, label = 'Tau Photoel.Abs')
    plt.plot(energie, tauNP, linewidth=1, label = 'Tau Nuclear.Pr')
    plt.plot(energie, tauEP, linewidth=1, label = 'Tau Elec.Pr')
    
    plt.legend()

    book.save("Données.xlsx")

#Cuivre........................................................................
def Affiche_Cuivre():
    
    book = op.load_workbook('Données.xlsx')
    
    #Variables
    Cuivre = book['Cuivre']
    
    #Listes
    energie = []
    tau = []
    tauPA = []
    tauNP = []
    tauEP = []
    
    for i in range(4,Cuivre.max_row+1):
        energie.append(float(Cuivre.cell(row=i, column=1).value))
        tau.append(float(Cuivre.cell(row=i, column=2).value))
        tauPA.append(float(Cuivre.cell(row=i, column=3).value))
        tauNP.append(float(Cuivre.cell(row=i, column=4).value))
        tauEP.append(float(Cuivre.cell(row=i, column=5).value))
    
    #Vérification
    print("\nListe tau : ", tau)
    print("\nListe tauPA : ", tauPA)
    print("\nListe tauNP : ", tauNP)
    print("\nListe tauEP : ", tauEP)
    
    plt.xlabel("Energie du photon (MEV)")
    plt.ylabel('Tau (cm2/g)')
    
    plt.xlim(0.001,1e+05)
    plt.ylim(1e-09,1e+04)
    
    plt.xscale('log')
    plt.yscale('log')
    
    plt.grid(True,which="both",linestyle='--')
    
    
    plt.plot(energie, tau, linewidth=1, label = 'Tau Inc.Scatter')
    plt.plot(energie, tauPA, linewidth=1, label = 'Tau Photoel.Abs')
    plt.plot(energie, tauNP, linewidth=1, label = 'Tau Nuclear.Pr')
    plt.plot(energie, tauEP, linewidth=1, label = 'Tau Elec.Pr')
    
    plt.legend()

    book.save("Données.xlsx")


# =============================================================================
# Programme principal
# =============================================================================

Affiche_Aluminium()
Affiche_Plomb()
Affiche_Cobalt()
Affiche_Cuivre()
